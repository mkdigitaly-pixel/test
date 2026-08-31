#!/usr/bin/env python3
"""Сбор номенклатуры MiniMed с minimed.ru и minimed.nt-rt.ru."""

from __future__ import annotations

import json
import re
import time
from collections import OrderedDict
from html import unescape
from pathlib import Path
from typing import Iterable
from urllib.parse import urljoin, urlparse

import requests
import xlrd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
CACHE = DATA / "cache"

BASE = "https://minimed.ru"
NTRT = "https://minimed.nt-rt.ru"
PRICE_XLS = BASE + "/price.xls"
CATALOG_PDF = BASE + "/catalog.pdf"
NTRT_PRICE_PDF = NTRT + "/images/showcase/pricelis.pdf"
NTRT_CATALOG_PDF = NTRT + "/images/showcase/katalogminimed.pdf"

UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)
DELAY = 0.45
VAT_DEFAULT = 22.0

HEADERS = [
    "Модель",
    "Наименование",
    "Категория",
    "Цена",
    "Артикул",
    "Ссылка на товар",
    "Источник",
]

ROOT_CATEGORIES = [
    "Лабораторная посуда и принадлежности из стекла",
    "Лабораторная посуда и принадлежности из пластика",
    "Лабораторная посуда и принадлежности из фарфора",
    "Лабораторная посуда и принадлежности прочие",
    "Лабораторное оборудование",
    "Пробирки вакуумные стерильные. Принадлежности для забора крови.",
    "Красители и химические реактивы",
    "Спецпредложение",
]

NTRT_SECTIONS = [
    (
        "Лабораторная посуда и принадлежности из стекла",
        NTRT + "/catalog/laboratornaya-posuda-i-prinadleznosti-iz-stekla",
    ),
    (
        "Лабораторная посуда и принадлежности из пластика",
        NTRT + "/catalog/laboratornaya-posuda-i-prinadleznosti-iz-plastika",
    ),
    (
        "Лабораторная посуда и принадлежности из фарфора",
        NTRT + "/catalog/laboratornaya-posuda-i-prinadleznosti-iz-farfora",
    ),
    (
        "Лабораторная посуда и принадлежности прочие",
        NTRT + "/catalog/laboratornaya-posuda-i-prinadleznosti-prochie",
    ),
    (
        "Лабораторное оборудование",
        NTRT + "/catalog/laboratornoe-oborudovanie",
    ),
    (
        "Пробирки вакуумные стерильные",
        NTRT + "/catalog/probirki-vakuumnye-sterilnye",
    ),
    (
        "Принадлежности для забора крови",
        NTRT + "/catalog/prinadleznosti-dlya-zabora-krovi",
    ),
    (
        "Красители и химические реактивы",
        NTRT + "/catalog/krasiteli-i-himiceskie-reaktivy",
    ),
]

# Первое слово группы (мн.ч.) → ед.ч. для наименования.
SINGULAR_FIRST = [
    (r"^Аквадистилляторы\b", "Аквадистиллятор"),
    (r"^Ареометры\b", "Ареометр"),
    (r"^Бани\b", "Баня"),
    (r"^Банки\b", "Банка"),
    (r"^Баллоны\b", "Баллон"),
    (r"^Бахилы\b", "Бахила"),
    (r"^Бинты\b", "Бинт"),
    (r"^Бутирометры\b", "Бутирометр"),
    (r"^Бутылки\b", "Бутылка"),
    (r"^Бутыли\b", "Бутыль"),
    (r"^Бюретки\b", "Бюретка"),
    (r"^Ванночки\b", "Ванночка"),
    (r"^Воронки\b", "Воронка"),
    (r"^Держатели\b", "Держатель"),
    (r"^Дозаторы\b", "Дозатор"),
    (r"^Емкости\b", "Емкость"),
    (r"^Ерши\b", "Ёрш"),
    (r"^Зажимы\b", "Зажим"),
    (r"^Иглы\b", "Игла"),
    (r"^Камеры\b", "Камера"),
    (r"^Капельницы\b", "Капельница"),
    (r"^Капилляры\b", "Капилляр"),
    (r"^Каплеуловители\b", "Каплеуловитель"),
    (r"^Кастрюли\b", "Кастрюля"),
    (r"^Ковши\b", "Ковш"),
    (r"^Колбы\b", "Колба"),
    (r"^Колпачки\b", "Колпачок"),
    (r"^Комплекты\b", "Комплект"),
    (r"^Контейнеры\b", "Контейнер"),
    (r"^Корзины\b", "Корзина"),
    (r"^Красители\b", "Краситель"),
    (r"^Кружки\b", "Кружка"),
    (r"^Кюветы\b", "Кювета"),
    (r"^Лампы\b", "Лампа"),
    (r"^Лодочки\b", "Лодочка"),
    (r"^Лопаточки\b", "Лопаточка"),
    (r"^Ложки\b", "Ложка"),
    (r"^Лотки\b", "Лоток"),
    (r"^Лупы\b", "Лупа"),
    (r"^Магниты\b", "Магнит"),
    (r"^Маски\b", "Маска"),
    (r"^Мензурки\b", "Мензурка"),
    (r"^Микроскопы\b", "Микроскоп"),
    (r"^Наконечники\b", "Наконечник"),
    (r"^Насадки\b", "Насадка"),
    (r"^Пакеты\b", "Пакет"),
    (r"^Палочки\b", "Палочка"),
    (r"^Переходы\b", "Переход"),
    (r"^Переходники\b", "Переходник"),
    (r"^Перчатки\b", "Перчатка"),
    (r"^Пестики\b", "Пестик"),
    (r"^Петли\b", "Петля"),
    (r"^Пикнометры\b", "Пикнометр"),
    (r"^Пинцеты\b", "Пинцет"),
    (r"^Пипетаторы\b", "Пипетатор"),
    (r"^Пипетки\b", "Пипетка"),
    (r"^Планшеты\b", "Планшет"),
    (r"^Пластины\b", "Пластина"),
    (r"^Плитки\b", "Плитка"),
    (r"^Полислайды\b", "Полислайд"),
    (r"^Приборы\b", "Прибор"),
    (r"^Пробирки\b", "Пробирка"),
    (r"^Пробки\b", "Пробка"),
    (r"^Промывалки\b", "Промывалка"),
    (r"^Растворы\b", "Раствор"),
    (r"^Салфетки\b", "Салфетка"),
    (r"^Скарификаторы\b", "Скарификатор"),
    (r"^Скальпели\b", "Скальпель"),
    (r"^Склянки-аспираторы\b", "Склянка-аспиратор"),
    (r"^Склянки\b", "Склянка"),
    (r"^Сосуды\b", "Сосуд"),
    (r"^Спиртовки\b", "Спиртовка"),
    (r"^Спринцовки\b", "Спринцовка"),
    (r"^Стаканчики\b", "Стаканчик"),
    (r"^Стаканы\b", "Стакан"),
    (r"^Ступки\b", "Ступка"),
    (r"^Счетчики\b", "Счетчик"),
    (r"^Секундомеры\b", "Секундомер"),
    (r"^Тампон-зонды\b", "Тампон-зонд"),
    (r"^Термометры\b", "Термометр"),
    (r"^Тигли\b", "Тигель"),
    (r"^Трубки\b", "Трубка"),
    (r"^Устройства\b", "Устройство"),
    (r"^Фильтры\b", "Фильтр"),
    (r"^Фитили\b", "Фитиль"),
    (r"^Флаконы\b", "Флакон"),
    (r"^Холодильники\b", "Холодильник"),
    (r"^Центрифуги\b", "Центрифуга"),
    (r"^Цилиндры\b", "Цилиндр"),
    (r"^Чаши\b", "Чаша"),
    (r"^Чашки\b", "Чашка"),
    (r"^Часы\b", "Часы"),
    (r"^Шапочки\b", "Шапочка"),
    (r"^Шпатели\b", "Шпатель"),
    (r"^Шприцы\b", "Шприц"),
    (r"^Штативы\b", "Штатив"),
    (r"^Щетки\b", "Щетка"),
    (r"^Щитки\b", "Щиток"),
    (r"^Эксикаторы\b", "Эксикатор"),
    (r"^Элементы\b", "Элемент"),
    (r"^Изгибы\b", "Изгиб"),
]

ADJ_PLURAL = [
    (r"\bлабораторные\b", "лабораторная"),
    (r"\bстеклянные\b", "стеклянная"),
    (r"\bпластиковые\b", "пластиковая"),
    (r"\bфарфоровые\b", "фарфоровая"),
    (r"\bвакуумные\b", "вакуумная"),
    (r"\bстерильные\b", "стерильная"),
    (r"\bнестерильные\b", "нестерильная"),
    (r"\bмедицинские\b", "медицинская"),
    (r"\bмерные\b", "мерная"),
    (r"\bградуированные\b", "градуированная"),
    (r"\bконические\b", "коническая"),
    (r"\bкруглодонные\b", "круглодонная"),
    (r"\bплоскодонные\b", "плоскодонная"),
    (r"\bделительные\b", "делительная"),
    (r"\bкерамические\b", "керамическая"),
    (r"\bметаллические\b", "металлическая"),
    (r"\bрезиновые\b", "резиновая"),
    (r"\bсиликоновые\b", "силиконовая"),
    (r"\bпоршневые\b", "поршневая"),
    (r"\bэлектрические\b", "электрическая"),
    (r"\bзащитные\b", "защитная"),
    (r"\bпроцедурные\b", "процедурная"),
    (r"\bпсихрометрические\b", "психрометрическая"),
    (r"\bгистологические\b", "гистологическая"),
    (r"\bмикробиологические\b", "микробиологическая"),
    (r"\bсерологические\b", "серологическая"),
    (r"\bбактериологические\b", "бактериологическая"),
    (r"\bинъекционные\b", "инъекционная"),
    (r"\bофисные\b", "офисная"),
    (r"\bфильтровальные\b", "фильтровальная"),
    (r"\bбуферные\b", "буферный"),
    (r"\bхимические\b", "химический"),
    (r"^Буферные растворы\b", "Буферный раствор"),
]


def cell(value) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"[\t\r\n]+", " ", unescape(text)).strip()


def norm_art(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, "g")
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def parse_money(raw) -> float | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw)
    text = text.replace("\xa0", " ").replace(" ", "").replace("руб.", "").replace("руб", "")
    text = text.replace(",", ".")
    text = re.sub(r"[^0-9.]", "", text)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def price_without_vat(gross, vat) -> str:
    amount = parse_money(gross)
    if amount is None or amount <= 0:
        return ""
    rate = 0.0
    if vat is None or vat == "":
        rate = VAT_DEFAULT
    elif isinstance(vat, (int, float)):
        rate = float(vat)
    else:
        text = str(vat).strip().lower().replace("%", "").replace(",", ".")
        if "без" in text:
            rate = 0.0
        else:
            m = re.search(r"(\d+(?:\.\d+)?)", text)
            rate = float(m.group(1)) if m else VAT_DEFAULT
    if rate > 0:
        amount = amount / (1.0 + rate / 100.0)
    value = int(round(amount))
    return str(value) if value > 0 else ""


def singular_type(group: str) -> str:
    name = cell(group)
    if not name:
        return ""
    for pat, repl in SINGULAR_FIRST:
        if re.search(pat, name, flags=re.I):
            name = re.sub(pat, repl, name, count=1, flags=re.I)
            break
    # согласовать первое прилагательное после существительного
    parts = name.split(" ", 1)
    if len(parts) == 2:
        head, rest = parts
        for pat, repl in ADJ_PLURAL:
            rest = re.sub(pat, repl, rest, count=1, flags=re.I)
        name = f"{head} {rest}"
    return name


MODEL_RES = [
    re.compile(r"\b(ПЗК-\d+[xх\*×]\d+)\b", re.I),
    re.compile(r"\b(\d{2}\.\d{3,}\.\d{2,})\b"),  # 20.1288.800, не ТУ 32.50.50
    re.compile(r"\b(?!ТУ\b)(?!ГОСТ\b)(?!РУ\b)(?!РЗН\b)([А-ЯA-ZЁ]{1,3}\s\d{3,}(?:\s?-\s?\d+)?)\b"),
    # ГОСТ 25336 и аналоги: К-1-50-29/32, П-2-50-22, В-1-25, Кн-1-50-14/23, КГУ-2-1-100-19/26-14/23
    re.compile(
        r"\b([А-ЯA-ZЁ][А-ЯA-ZЁа-яa-zё]{0,5}(?:-[А-ЯA-ZЁа-яa-zё])?-?\d[\d./xх\*×\-]*"
        r"(?:-[А-ЯA-ZЁа-яa-zё0-9][\d./xх\*×\-]*)*)\b"
    ),
    re.compile(r"\b(\d(?:-\d+){3,5}(?:,\d+)?)\b"),  # ГОСТ 1-3-2-10-0,05
    # мерные / Къельдаля / Бунзена: 2а-25-2, 1-50-14/23, 2-100-22
    re.compile(r"(?<!\d)(\d[аaАA]-\d+(?:-\d+)+)\b"),
    re.compile(r"(?<!\d)(\d(?:-\d+){1,3}(?:/\d+)?(?:-\d+(?:/\d+)?)*)\b"),
]

VOLUME_ONLY_RE = re.compile(
    r"^(?:\d+(?:[.,]\d+)?\s*(?:мкл|мл|л|мм|см|г|кг|шт)\b.*)$",
    re.I,
)


def _norm_dashes(text: str) -> str:
    return cell(text).replace("–", "-").replace("—", "-").replace("−", "-")


def _norm_size(text: str) -> str:
    t = _norm_dashes(text)
    return cell(t.replace("×", "x").replace("*", "x").replace("х", "x"))


def _is_plausible_model(model: str) -> bool:
    t = cell(model)
    if len(t) < 3 or t.endswith(","):
        return False
    up = t.upper()
    if up.startswith(("ТУ", "ГОСТ", "ОСТ", "ISO", "РЗН", "ИНН", "РУ ")):
        return False
    if re.fullmatch(r"\d{1,3}/\d{1,4}", t):
        return False
    # номер ТУ 9464-019-29508133-2015, не типоразмер
    if re.fullmatch(r"\d+(?:-\d+){2,}", t) and len(re.sub(r"\D", "", t)) >= 12:
        return False
    return True


def _first_model(text: str) -> str:
    blob = _norm_dashes(text)
    for rx in MODEL_RES:
        m = rx.search(blob)
        if m:
            model = _norm_size(m.group(1))
            if _is_plausible_model(model):
                return model
    return ""


def _xls_head_model(raw_name: str) -> str:
    """В прайсе модель стоит первым полем: «К-1-50-29/32; уп. 12/192 шт.»."""
    blob = _norm_dashes(raw_name)
    if not blob:
        return ""
    main = blob.split(";")[0].strip()
    head = main.split(",")[0].strip()
    if VOLUME_ONLY_RE.match(head):
        return ""
    if re.fullmatch(
        r"[А-ЯA-ZЁа-яёa-z]{1,8}(?:-[А-ЯA-ZЁа-яёa-z])?-?\d[\d./xх\*×\-]*",
        head,
        flags=re.I,
    ):
        return _norm_size(head) if _is_plausible_model(head) else ""
    if re.fullmatch(r"\d[аaАA]?(?:-\d+)+(?:/\d+)?(?:-\d+(?:/\d+)?)*", head):
        return _norm_size(head) if _is_plausible_model(head) else ""
    if re.fullmatch(r"[А-ЯA-ZЁ]{1,4}\s+\d[\d./\-]*", head):
        return _norm_size(head) if _is_plausible_model(head) else ""
    return ""


def extract_model(raw_name: str, site_name: str = "") -> str:
    head = _xls_head_model(raw_name)
    if head:
        return head
    # прайс надёжнее карточки: в TSV модель уже могла быть вырезана из названия
    for source in (raw_name, site_name):
        found = _first_model(source)
        if found:
            return found
    blob = _norm_dashes(raw_name)
    main = blob.split(";")[0].strip()
    main = re.sub(r"\([^)]*\)", " ", main)
    main = re.sub(r"\s+", " ", main).strip()
    if VOLUME_ONLY_RE.match(main):
        cleaned = re.sub(r",?\s*спецзаказ\s*$", "", main, flags=re.I).strip(" ,;")
        cleaned = _norm_size(cleaned)
        if re.search(r"\d+\s*x\s*\d+", cleaned, re.I):
            return cell(cleaned)
        m = re.match(r"(\d+(?:[.,]\d+)?\s*(?:мкл|мл|л))\b", cleaned, re.I)
        if m:
            return cell(m.group(1))
        return ""
    return ""


def strip_model(name: str, model: str) -> str:
    text = cell(name)
    if not text:
        return ""
    text = re.sub(r"\bминимед\w*\b", " ", text, flags=re.I)
    text = re.sub(r"\bminimed\w*\b", " ", text, flags=re.I)
    if model:
        variants = {model, model.replace("x", "х"), model.replace("х", "x")}
        spaced = re.sub(r"\s*-\s*", " -", model)
        variants.add(spaced)
        for var in variants:
            text = re.sub(re.escape(var), " ", text, flags=re.I)
        # хвост диапазона из модели (АУ 1000-1050 → убрать 1000-1050)
        tail = re.search(r"(\d+\s*-\s*\d+)\s*$", model)
        if tail:
            text = re.sub(re.escape(tail.group(1)), " ", text)
    text = re.sub(r"\s+,", ",", text)
    text = re.sub(r"\s{2,}", " ", text)
    text = re.sub(r"^[\s,;./\-]+", "", text)
    text = re.sub(r"[\s,;]+$", "", text)
    # убрать пустые скобки
    text = re.sub(r"\(\s*\)", "", text)
    text = re.sub(r"\s{2,}", " ", text).strip(" ,;")
    return cell(text)


def _singular_shape(word: str) -> str:
    w = cell(word)
    if not w:
        return ""
    for pat, repl in ADJ_PLURAL:
        nxt = re.sub(pat, repl, w, flags=re.I)
        if nxt != w:
            return nxt
    if re.search(r"(ые|ие)$", w, flags=re.I) and w[0].islower():
        return re.sub(r"ые$", "ая", re.sub(r"ие$", "яя", w, flags=re.I), flags=re.I)
    return w


def name_from_group(group: str) -> str:
    """Колбы лабораторные (круглодонные: исполнение 1 - со шлифом, тип К)
    → Колба круглодонная со шлифом."""
    text = cell(group)
    if "/" in text:
        text = text.rsplit("/", 1)[-1]
    m = re.match(
        r"^([А-ЯЁа-яёA-Za-z][А-ЯЁа-яёA-Za-z\-]*(?:\s+[А-ЯЁа-яёA-Za-z\-]+)*)\s*(?:\(([^)]*)\))?(.*)$",
        text,
    )
    if not m:
        return singular_type(text)
    head = m.group(1).strip()
    inside = (m.group(2) or "").strip()
    after = (m.group(3) or "").strip()
    noun = singular_type(head).split()[0] if singular_type(head) else head
    if not inside:
        body = singular_type(head)
        return cell(f"{body} {after}".strip())
    if ":" in inside:
        shape, rest = inside.split(":", 1)
    else:
        shape, rest = inside, ""
    shape = shape.strip()
    rest = rest.strip()
    rest = re.sub(r"исполнение\s*\d+\s*[аaАA]?\s*-?\s*", "", rest, flags=re.I)
    rest = re.sub(r",?\s*тип\s+[А-ЯA-ZЁа-яё0-9\-]+\s*", "", rest, flags=re.I)
    rest = re.sub(r"\s{2,}", " ", rest).strip(" ,;-")
    bits = [noun]
    if shape:
        # имя собственное (Къельдаля, Бунзена, Энглера) — как есть
        if shape[0].isupper() and not re.search(r"(ые|ие|ый|ой|ая)$", shape, flags=re.I):
            bits.append(shape)
        else:
            shaped = _singular_shape(shape)
            if shaped.lower() not in {"лабораторная", "лабораторные"}:
                bits.append(shaped)
    if rest:
        bits.append(rest)
    if after:
        bits.append(after.strip(" ,;"))
    return cell(" ".join(b for b in bits if b))


def _peel_pack_marks(name: str) -> tuple[str, str]:
    text = cell(name)
    pack = ""
    m = re.search(r"(?:,\s*)?(уп\.?\s*.+)$", text, flags=re.I)
    if m:
        pack = m.group(0).strip(" ,;")
        text = text[: m.start()].strip(" ,;")
    marks: list[str] = []
    for rx in (
        r"\b[бс]/дел\.?",
        r"\bс\s*дел(?:ением|ениями)?\.?",
        r"\bбез\s+делений\b",
        r"\bб/дел\.?",
        r"\bТС\b",
    ):
        mm = re.search(rx, text, flags=re.I)
        if mm:
            marks.append(mm.group(0).strip())
            text = re.sub(rx, " ", text, flags=re.I)
    text = re.sub(r"\b(ТУ|ГОСТ|ОСТ)\s*[\d.\-/]+", " ", text, flags=re.I)
    text = re.sub(r"[,;()]+", " ", text)
    text = re.sub(r"\s{2,}", " ", text).strip(" ,;")
    seen: set[str] = set()
    tail: list[str] = []
    for item in marks:
        key = item.lower()
        if key not in seen:
            seen.add(key)
            tail.append(item)
    if pack:
        tail.append(pack)
    return text, ", ".join(tail)


def _group_body(group: str) -> str:
    if re.search(r"исполнение", group or "", flags=re.I):
        return name_from_group(group) or singular_type(group)
    return singular_type(group)


def _should_use_group_name(name: str, group: str) -> bool:
    leaf = group.rsplit("/", 1)[-1] if group else ""
    if not re.search(r"исполнение", leaf, flags=re.I):
        return False
    core, _ = _peel_pack_marks(name)
    words = [w for w in core.split() if w]
    if not words:
        return True
    group_name = name_from_group(group)
    group_words = {w.lower() for w in group_name.split() if w}
    name_words = {w.lower() for w in words}
    if name_words and name_words <= group_words:
        return True
    if re.search(r"\b(ТУ|ГОСТ|ОСТ)\b", name, flags=re.I) and len(words) <= 2:
        return True
    return False


def build_name(site_name: str, xls_name: str, group: str, model: str) -> str:
    if site_name:
        name = strip_model(site_name, model)
    else:
        xls = cell(xls_name)
        main, *rest = [p.strip() for p in xls.split(";") if p.strip()]
        pack = ""
        extra = []
        for part in rest:
            if part.lower().startswith("уп"):
                pack = part
            else:
                extra.append(part)
        if VOLUME_ONLY_RE.match(main) or (model and main.replace(" ", "") == model.replace(" ", "")):
            body = _group_body(group)
            bits = [body]
            if VOLUME_ONLY_RE.match(main):
                bits.append(main)
            elif extra:
                bits.extend(extra)
            elif main and not model:
                bits.append(main)
            name = ", ".join(b for b in bits if b)
            if pack:
                name = f"{name}, {pack}"
        else:
            # описание в прайсе: «для молока» в скобках
            desc = ""
            m = re.search(r"\(([^)]+)\)", xls)
            if m:
                desc = m.group(1).strip()
            body = _group_body(group)
            bits = [body]
            leftover = strip_model(main, model)
            leftover = leftover.strip(" ;,")
            if leftover and leftover.lower() not in body.lower():
                bits.append(leftover)
            if desc and desc.lower() not in " ".join(bits).lower():
                bits.append(desc)
            name = ", ".join(b for b in bits if b)
            if pack:
                name = f"{name}, {pack}"
        name = strip_model(name, model)
    name = re.sub(r"\s{2,}", " ", name).strip(" ,;")
    name = re.sub(r",\s*,+", ",", name)
    name = re.sub(r"\s{2,}", " ", name).strip(" ,;")
    # если название начинается с кода — вынести существительное вперёд
    if name and not re.match(r"[А-ЯЁ]", name):
        m = re.search(r"([А-ЯЁ][А-ЯЁа-яё\-]*(?:\s+\S+){0,12})", name)
        if m:
            head = m.group(1).rstrip(" :;,")
            rest = (name[: m.start()] + " " + name[m.end() :]).strip(" :;,")
            name = cell(f"{head} {rest}".strip())
    if name and name[0].islower():
        name = name[0].upper() + name[1:]
    # ед.ч. первого слова
    if name:
        name = singular_type(name)
    if group and _should_use_group_name(name, group):
        body = name_from_group(group)
        if body:
            _, tail = _peel_pack_marks(name)
            name = body
            if tail:
                name = f"{name}, {tail}"
    name = re.sub(r"\s{2,}", " ", name).strip(" ,;")
    name = re.sub(r",\s*,+", ",", name)
    return name


def is_root_category(name: str) -> bool:
    n = cell(name).rstrip(".")
    for root in ROOT_CATEGORIES:
        if n == root.rstrip(".") or name.strip() == root:
            return True
        if name.strip().startswith("Пробирки вакуумные стерильные"):
            return True
    return False


class Http:
    def __init__(self):
        self.s = requests.Session()
        self.s.headers.update({"User-Agent": UA, "Accept-Language": "ru,en;q=0.8"})

    def get(self, url: str, retries: int = 4) -> requests.Response:
        last = None
        for i in range(retries):
            try:
                time.sleep(DELAY)
                r = self.s.get(url, timeout=60)
                if r.status_code in (429, 500, 502, 503, 504):
                    time.sleep(2 * (i + 1))
                    last = r
                    continue
                r.raise_for_status()
                return r
            except requests.RequestException as exc:
                last = exc
                time.sleep(2 * (i + 1))
        if isinstance(last, requests.Response):
            last.raise_for_status()
        raise last  # type: ignore

    def download(self, url: str, dest: Path) -> Path:
        dest.parent.mkdir(parents=True, exist_ok=True)
        if dest.exists() and dest.stat().st_size > 1000:
            return dest
        print(f"  download {url}")
        r = self.get(url)
        dest.write_bytes(r.content)
        return dest


def parse_catalog_tree(html: str) -> tuple[list[tuple[str, str]], list[tuple[str, str, str]]]:
    """Корни и листья (root_name, leaf_name, url)."""
    soup = BeautifulSoup(html, "lxml")
    roots: list[tuple[str, str]] = []
    leaves: list[tuple[str, str, str]] = []
    # боковое меню каталога
    for root_li in soup.select("ul li"):
        pass
    # надёжнее: все ссылки /catalog/x/ и /catalog/x/y/
    seen = set()
    root_map: OrderedDict[str, str] = OrderedDict()
    for a in soup.select('a[href^="/catalog/"]'):
        href = a.get("href", "").split("?")[0]
        if not href.endswith("/"):
            href += "/"
        if any(x in href for x in ("/tag/", "/discount/", "/lider/", "/new/", "/sale/")):
            continue
        parts = [p for p in href.strip("/").split("/") if p]
        if len(parts) < 2:
            continue
        name = cell(a.get_text())
        if not name or name.lower() in {"каталог", "подробнее"}:
            continue
        if href in seen:
            continue
        seen.add(href)
        url = urljoin(BASE, href)
        if len(parts) == 2:
            root_map[parts[1]] = name
            roots.append((name, url))
        elif len(parts) >= 3:
            root_name = root_map.get(parts[1], "")
            leaves.append((root_name, name, url))
    return roots, leaves


def pager_pages(html: str, page_url: str) -> list[str]:
    pages = {page_url}
    for m in re.finditer(r"PAGEN_1=(\d+)", html):
        n = int(m.group(1))
        if n > 1:
            sep = "&" if "?" in page_url else "?"
            # drop existing PAGEN
            base = re.sub(r"([?&])PAGEN_1=\d+", "", page_url).rstrip("&?")
            sep = "&" if "?" in base else "?"
            pages.add(f"{base}{sep}PAGEN_1={n}")
    # sequential 1..max
    nums = [1]
    for u in list(pages):
        m = re.search(r"PAGEN_1=(\d+)", u)
        if m:
            nums.append(int(m.group(1)))
    max_n = max(nums)
    out = []
    base = re.sub(r"([?&])PAGEN_1=\d+", "", page_url).rstrip("&?")
    for n in range(1, max_n + 1):
        if n == 1:
            out.append(base)
        else:
            sep = "&" if "?" in base else "?"
            out.append(f"{base}{sep}PAGEN_1={n}")
    return out


def parse_listing(html: str) -> list[dict]:
    soup = BeautifulSoup(html, "lxml")
    items = []
    for card in soup.select(".product_card"):
        title_a = card.select_one(".product__title a, a.s_name")
        if not title_a or not title_a.get("href"):
            continue
        href = urljoin(BASE, title_a["href"])
        name = cell(title_a.get("title") or title_a.get_text())
        art = ""
        art_el = card.select_one(".product__article")
        if art_el:
            art = norm_art(re.sub(r"артикул:\s*", "", art_el.get_text(" ", strip=True), flags=re.I))
        price = None
        val = card.select_one(".product__price .value")
        if val:
            price = parse_money(val.get_text())
        items.append({"name": name, "url": href, "article": art, "price": price})
    if items:
        return items
    # JSON-LD fallback
    for script in soup.select('script[type="application/ld+json"]'):
        try:
            data = json.loads(script.string or "")
        except Exception:
            continue
        stack = [data]
        while stack:
            node = stack.pop()
            if isinstance(node, list):
                stack.extend(node)
                continue
            if not isinstance(node, dict):
                continue
            stack.extend(node.values())
            if node.get("@type") == "Product" and node.get("url"):
                offers = node.get("offers") or {}
                if isinstance(offers, list):
                    offers = offers[0] if offers else {}
                items.append(
                    {
                        "name": cell(node.get("name")),
                        "url": urljoin(BASE, str(node.get("url"))),
                        "article": norm_art(node.get("mpn") or node.get("sku") or ""),
                        "price": parse_money(offers.get("price")),
                    }
                )
    return items


def crawl_site(http: Http) -> dict[str, dict]:
    """article -> {name, url, price, category}."""
    print("Каталог minimed.ru …")
    html = http.get(BASE + "/catalog/").text
    roots, leaves = parse_catalog_tree(html)
    print(f"  корней {len(roots)}, листьев {len(leaves)}")
    by_art: dict[str, dict] = {}
    for i, (root_name, leaf_name, url) in enumerate(leaves, 1):
        category = "/".join(p for p in (root_name, leaf_name) if p)
        try:
            page1 = http.get(url).text
        except Exception as exc:
            print(f"  skip {url}: {exc}")
            continue
        pages = pager_pages(page1, url)
        htmls = [page1]
        for p in pages[1:]:
            try:
                htmls.append(http.get(p).text)
            except Exception as exc:
                print(f"  skip {p}: {exc}")
        for h in htmls:
            for item in parse_listing(h):
                art = item["article"]
                if not art:
                    continue
                rec = {
                    "name": item["name"],
                    "url": item["url"],
                    "price": item["price"],
                    "category": category,
                }
                if art not in by_art:
                    by_art[art] = rec
        if i == 1 or i % 20 == 0 or i == len(leaves):
            print(f"  [{i}/{len(leaves)}] {leaf_name}: карточек {len(by_art)}")
    print(f"  итого карточек: {len(by_art)}")
    return by_art


def parse_price_xls(path: Path) -> list[dict]:
    wb = xlrd.open_workbook(str(path))
    sh = wb.sheet_by_index(0)
    rows = []
    root = ""
    group = ""
    for r in range(5, sh.nrows):
        art = norm_art(sh.cell_value(r, 0))
        name = cell(sh.cell_value(r, 1))
        if not art and not name:
            continue
        if not art:
            if is_root_category(name):
                root = name
                group = ""
            else:
                group = name
            continue
        vat = sh.cell_value(r, 2)
        wholesale = sh.cell_value(r, 5)
        cat = "/".join(p for p in (root, group) if p)
        rows.append(
            {
                "article": art,
                "xls_name": name,
                "root": root,
                "group": group,
                "category": cat,
                "vat": vat,
                "wholesale": wholesale,
            }
        )
    return rows


def pdf_articles(path: Path) -> set[str]:
    arts: set[str] = set()
    try:
        import pdfplumber
    except ImportError:
        return arts
    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            arts.update(re.findall(r"\b(\d{8})\b", text))
    return arts


def write_tsv(path: Path, rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        f.write("\t".join(HEADERS) + "\n")
        for row in rows:
            if not row:
                f.write("\n")
            else:
                f.write("\t".join(row) + "\n")


def write_xlsx(path: Path, rows: list[list[str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Номенклатура"
    ws.append(HEADERS)
    for cell_ in ws[1]:
        cell_.font = Font(bold=True)
    for row in rows:
        ws.append(row if row else [""] * 7)
    ws.auto_filter.ref = f"A1:G{ws.max_row}"
    ws.freeze_panes = "A2"
    wb.save(path)


def ntrt_fallback(category: str) -> str:
    cat = category.split("/")[0] if category else ""
    for name, url in NTRT_SECTIONS:
        if cat.startswith(name.split(".")[0]) or name.startswith(cat.split(".")[0]):
            return url
    return NTRT + "/catalog"


def assemble(
    xls_rows: list[dict],
    site: dict[str, dict],
    extra_arts: set[str],
    pdf_products: list[dict] | None = None,
) -> list[list[str]]:
    pdf_products = pdf_products or []
    used = set()
    out: list[list[str]] = []
    prev_cat = None
    first = True

    def emit_group(cat: str):
        nonlocal first, prev_cat
        if prev_cat == cat:
            return
        if not first:
            out.append([])  # пустая строка-разделитель категорий
        out.append([cat, "", "", "", "", "", ""])
        prev_cat = cat
        first = False

    for rec in xls_rows:
        art = rec["article"]
        site_rec = site.get(art) or {}
        model = extract_model(rec["xls_name"], site_rec.get("name", ""))
        name = build_name(site_rec.get("name", ""), rec["xls_name"], rec["group"] or rec["root"], model)
        cat = rec["category"] or site_rec.get("category") or rec["root"]
        price = price_without_vat(rec["wholesale"] or site_rec.get("price"), rec["vat"])
        url = site_rec.get("url") or ""
        if url:
            source = "minimed.ru"
        else:
            url = ntrt_fallback(cat)
            source = "minimed.ru (прайс)"
        emit_group(cat)
        out.append([cell(model), cell(name), cell(cat), price, art, url, source])
        used.add(art)

    # товары сайта, которых нет в прайсе
    extras = []
    for art, rec in site.items():
        if art in used:
            continue
        extras.append((rec.get("category") or "Прочее", rec, art))
    extras.sort(key=lambda x: (x[0], x[2]))
    for cat, rec, art in extras:
        model = extract_model("", rec.get("name", ""))
        name = build_name(rec.get("name", ""), "", cat.split("/")[-1], model)
        price = price_without_vat(rec.get("price"), VAT_DEFAULT)
        emit_group(cat)
        out.append(
            [
                cell(model),
                cell(name),
                cell(cat),
                price,
                art,
                rec.get("url", ""),
                "minimed.ru",
            ]
        )
        used.add(art)

    # типоразмеры из catalog.pdf, которых нет в прайсе/на сайте
    pdf_by_cat: dict[str, list] = {}
    for rec in pdf_products:
        art = rec.get("article") or ""
        if not art or art in used or art + "А" in used or art + "A" in used:
            continue
        name = rec.get("name") or ""
        if name == "Технические характеристики пипеток":
            rec = dict(rec)
            rec["name"] = "Пипетка градуированная"
            name = rec["name"]
        low = name.lower()
        if any(
            x in low
            for x in (
                "приобрета",
                "изготовлен",
                "поставля",
                "применяет",
                "в состав",
                "толщина",
                "предназнач",
                "диаметр гнезд",
                "наружный диаметр",
            )
        ) or name.lower().startswith(("диаметр", "наружный", "высота ", "ширина")):
            continue
        pdf_by_cat.setdefault(rec.get("category") or "PDF-каталог", []).append(rec)
        used.add(art)
    for cat, items in pdf_by_cat.items():
        emit_group(cat)
        for rec in items:
            out.append(
                [
                    cell(rec.get("model")),
                    cell(rec.get("name")),
                    cell(cat),
                    "",
                    rec["article"],
                    rec.get("url") or CATALOG_PDF,
                    rec.get("source") or "minimed.ru (PDF-каталог)",
                ]
            )
    return out


def pdf_from_tsv(path: Path) -> list[dict]:
    """Сохранить строки PDF-каталога при офлайн-пересборке без catalog.pdf."""
    out: list[dict] = []
    if not path.exists():
        return out
    for line in path.read_text(encoding="utf-8").splitlines()[1:]:
        if not line.strip():
            continue
        cols = line.split("\t")
        if len(cols) < 7:
            continue
        model, name, cat, _price, art, url, src = cols[:7]
        if not art or "PDF" not in src:
            continue
        out.append(
            {
                "model": model,
                "name": name,
                "category": cat,
                "article": art,
                "url": url,
                "source": src,
            }
        )
    return out


def site_from_tsv(path: Path) -> dict[str, dict]:
    site: dict[str, dict] = {}
    if not path.exists():
        return site
    for line in path.read_text(encoding="utf-8").splitlines()[1:]:
        if not line.strip():
            continue
        cols = line.split("\t")
        if len(cols) < 7:
            continue
        _model, name, cat, _price, art, url, src = cols[:7]
        if not art or "прайс" in src or "PDF" in src:
            continue
        site[art] = {"name": name, "url": url, "price": None, "category": cat}
    return site


def main() -> None:
    import sys

    DATA.mkdir(parents=True, exist_ok=True)
    CACHE.mkdir(parents=True, exist_ok=True)
    http = Http()
    offline = "--offline" in sys.argv

    xls_path = CACHE / "price.xls"
    if not xls_path.exists():
        xls_path = http.download(PRICE_XLS, xls_path)
    print("Разбор price.xls …")
    xls_rows = parse_price_xls(xls_path)
    print(f"  позиций в прайсе: {len(xls_rows)}")

    site_cache = CACHE / "site.json"
    tsv_path = DATA / "nomenclature.tsv"
    if offline and tsv_path.exists():
        site = site_from_tsv(tsv_path)
        print(f"  карточки из TSV: {len(site)}")
    elif site_cache.exists() and offline:
        site = json.loads(site_cache.read_text(encoding="utf-8"))
        print(f"  карточки из кэша: {len(site)}")
    else:
        site = crawl_site(http)
        site_cache.write_text(json.dumps(site, ensure_ascii=False), encoding="utf-8")
        print(f"  карточек на сайте: {len(site)}")

    extra_arts: set[str] = set()
    pdf_products: list[dict] = []
    catalog_path = CACHE / "catalog.pdf"
    if catalog_path.exists() or not offline:
        try:
            if not catalog_path.exists():
                catalog_path = http.download(CATALOG_PDF, catalog_path)
            from pdf_catalog import parse_catalog_pdf

            print("Разбор catalog.pdf …")
            pdf_products = parse_catalog_pdf(catalog_path, CATALOG_PDF)
            print(f"  позиций в PDF-каталоге: {len(pdf_products)}")
        except Exception as exc:
            print(f"  catalog.pdf: {exc}")
    if not pdf_products and offline:
        pdf_products = pdf_from_tsv(tsv_path)
        if pdf_products:
            print(f"  позиции PDF из TSV: {len(pdf_products)}")
    if not offline:
        for url, name in (
            (NTRT_CATALOG_PDF, "ntrt-katalog.pdf"),
            (NTRT_PRICE_PDF, "ntrt-pricelis.pdf"),
        ):
            try:
                path = http.download(url, CACHE / name)
                arts = pdf_articles(path)
                print(f"  {name}: артикулов {len(arts)}")
                extra_arts |= arts
            except Exception as exc:
                print(f"  {name}: {exc}")

    rows = assemble(xls_rows, site, extra_arts, pdf_products)
    goods = [r for r in rows if r and len(r) > 4 and r[4]]
    tsv = DATA / "nomenclature.tsv"
    xlsx = DATA / "nomenclature.xlsx"
    write_tsv(tsv, rows)
    write_xlsx(xlsx, rows)
    print(f"Готово: {len(goods)} товаров → {tsv}")


if __name__ == "__main__":
    main()
